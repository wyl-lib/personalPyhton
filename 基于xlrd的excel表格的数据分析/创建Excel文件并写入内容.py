import xlrd
import xlwt
import openpyxl

global nrows

def set_style(name,height,bold=False):

    style = xlwt.XFStyle()

    font = xlwt.Font()
    font.name = name
    font.bold = bold
    font.color_index = 4
    font.height = height

    style.font = font

    return style

def read_excel(file_name):
    '''
        读取表格file_datefrom_excel的数据(使用绝对路径)，遍历“分类号”，
        将相同的归为一行数据，整合不同内容的字段，如：“资产编号”使用下拉框进行数据存储，以及数据展示。
    '''
    global file_datefrom_excel
    data = xlrd.open_workbook(file_name)#文件名以及路径，如果路径或者文件名有中文给前面加一个r拜师原生字符.
    table = data.sheet_by_name("二教104")
    
    colum0 = []
    date_moveline = 1
    common_count = 0
    row0 = table.row_values(0, start_colx=0, end_colx=None)#返回由该行中所有单元格的数据组成的列表
    print("row0 {}".format(row0))

    ncols = table.ncols #获取列表的有效列数
    nrows = table.nrows  #获取该sheet中的有效行数
    '''
    print(sheet1.cell(1,0).value)#获取表格里的内容，三种方式
    print(sheet1.cell_value(1,0))
    print(sheet1.row(1)[0].value)
    '''
    for i in range(2,nrows):
        print(i)
        before = str(table.cell(i-1,3))
        now = str(table.cell(i,3))
        #print("下一个记录 {}".format(before))
        #print("当前记录 {}".format(now))
        #print(type(before))
        if(i==nrows-1):
            #返回由该行中所有单元格的数据组成的列表
            colum0.append(table.row_values(i, start_colx=0, end_colx=None))
            print(common_count)
            write_excel(common_count+1,nrows,ncols,row0,colum0)
            
        if(before==now):
            continue
        else:
            #返回由该行中所有单元格的数据组成的列表
            colum0.append(table.row_values(i, start_colx=0, end_colx=None))
            #print("这项记录共有:{}条".format(common_count))
            common_count += 1    
    
    #print(table.row(1)) #返回由该“行”中所有的单元格对象组成的列表
    #print(table.col_slice(1, start_rowx=0, end_rowx=None)) #返回由该列中所有的单元格对象组成的列表

def combox_excel():

    '''
        “资产编号”使用下拉框进行数据存储，进而用以数据展示。
    '''


def write_excel(line,nrows,ncols,row0,colum0):
    
    print("******写入列值 {}******".format(line)) 
    #创建一个Workbook 设置编码
    workbook = xlwt.Workbook()

    #创建一个Worksheet
    worksheet = workbook.add_sheet('My Worksheet',cell_overwrite_ok=True)

    #写第一行
    for i in range(0,len(row0)):
        worksheet.write(0,i,row0[i],set_style('Times New Roman',220,True))
    
    #写入列值
    for i in range(1,line):
        for j in range(0,ncols):
            worksheet.write(i,j,colum0[i][j],set_style('Times New Roman',220,False))
    '''
    write_merge(x, x + m, y, w + n, string, sytle)
    x表示行，y表示列，m表示跨行个数，n表示跨列个数，string表示要写入的单元格内容，
    style表示单元格样式。其中，x，y，w，h，都是以0开始计算的。
    这个和xlrd中的读合并单元格的不太一样。

    row0 = ["审核状态","资产编号","资产名称","分类号","分类名称","项目号","单价","总造价","套(件数)","计量单位",
    "使(领)用人","使用方向","使用单位","存放地点","单位性质","现状","生产厂家","出厂编号","型号","规格","经费来源",
    "资产来源","购置日期","入帐日期","调转入日期","档案编号","凭证号","验收单号","记帐类型","学科","学科类别",
    "归口审核人","归口审核单位","归口审核日期","归口审核意见","财务审核人","财务审核日期","财务审核意见"]
    colum0 = ["张三","李四","恋习Python","小明","小红","无名"]

    #参数对应：row，line，values   合并单元格函数
    worksheet.write(1,3,'2006/12/12')
    worksheet.write_merge(6,6,1,3,'行合并，未知')#合并“行”单元格
    worksheet.write_merge(1,2,3,3,'列合并，打游戏')#合并“列”单元格
    worksheet.write_merge(4,5,3,3,'打篮球')
    '''
    #保存文件
    workbook.save('Excel_create_test3.xls')
    print("@@@@@@@@@@@写入完成共_{}_记录@@@@@@@@@@".format(line))
    

def read_with_dropdown(book_name, sheet_name):
    #读取excel.xlsx格式文件
    wb = openpyxl.load_workbook(book_name)
    #读取sheet表
    ws = wb[sheet_name]
    # 获取内容存在下拉选的框数据
    validations = ws.data_validations.dataValidation
    #遍历存在下拉选的单元格
    for validation in validations:
        cell = validation.sqref
        result = validation.formula1
        print("单元格位置:"+str(cell)+",下拉选内容："+result)


if __name__ == '__main__':

    #read_with_dropdown()
    #"D:\WriteCode\PyhtonCode\基于xlrd的excel表格的数据分析\测试表格Test1.xlsx", "test2"
    file_datefrom_excel = "D:\WriteCode\PyhtonCode\基于xlrd的excel表格的数据分析\二教104.xls"
    read_excel(file_datefrom_excel)







